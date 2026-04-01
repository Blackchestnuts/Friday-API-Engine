from core.logger import logger
import yaml
import json
import os
import requests
import openpyxl
import allure

class FridayEngine:
    """Mark2 核心引擎：读取配置与表格，自动生成并执行接口请求"""
    def __init__(self, config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)
        self.systems = self.config['systems']
        self.tokens = {}
        
        # 中英文翻译官字典
        self.header_map = {
            "所属系统": "system_key", "用例名称": "case_name", "请求方式": "method",
            "接口路径": "path", "URL参数": "query_params", "请求体": "body_data",
            "提取变量": "extract_vars", "断言状态码": "assert_status",
            "断言路径": "assert_jsonpath", "预期结果": "assert_expected_value"
        }

    def _parse_json_path(self, data, path):
        if not path or not data: return None
        keys = path.lstrip('$.').split('.')
        temp = data
        for key in keys:
            if isinstance(temp, dict) and key in temp: temp = temp[key]
            else: return None
        return temp

    def _get_system_token(self, system_key):
        if system_key in self.tokens: return self.tokens[system_key]
        
        sys_conf = self.systems[system_key]
        auth_conf = sys_conf['auth']
        url = sys_conf['base_url'] + auth_conf['login_url']
        headers = sys_conf['headers']
        req_params = auth_conf.get('query_params')
        
        logger.info("🔐 [引擎鉴权] 正在向 %s 申请通行证...", system_key)
        response = requests.post(url, headers=headers, params=req_params, json=auth_conf['payload'])
        res_json = response.json()
        
        token = self._parse_json_path(res_json, auth_conf['token_extract_path'])
        if token:
            self.tokens[system_key] = token
            logger.info("🔑 [引擎鉴权] 通行证获取成功: %s...", token[:20])
            return token
        else:
            logger.error("⚠️ [引擎警报] 鉴权失败！返回内容: %s", res_json)
            return None
    def run_excel(self, excel_path):
        if not os.path.exists(excel_path):
            print(f"❌ 找不到弹药库文件: {excel_path}")
            return 0, 0

        print(f"📂 [引擎启动] 正在装载弹药库: {excel_path}")
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        passed, failed = 0, 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_case = dict(zip(headers, row))
            case = {}
            for cn_key, en_key in self.header_map.items():
                if cn_key in raw_case: case[en_key] = raw_case[cn_key]
            for key, val in raw_case.items():
                if key not in self.header_map: case[key] = val
            
            if not case.get('case_name'): continue
            case_name = case['case_name']
            
            try:
                with allure.step(case_name):
                    sys_key = case['system_key']
                    sys_conf = self.systems[sys_key]
                    base_url = sys_conf['base_url']
                    
                    req_headers = sys_conf['headers'].copy()
                    token = self._get_system_token(sys_key)
                    if token: req_headers[sys_conf['auth']['token_header_name']] = token
                    else: raise Exception("无通行证，终止执行")

                    url = base_url + case['path']
                    method = case['method'].upper()
                    params = json.loads(case['query_params']) if case.get('query_params') else None
                    body = json.loads(case['body_data']) if case.get('body_data') else None
                    
                    req_detail = f"Method: {method}\nURL: {url}\nHeaders: {json.dumps(req_headers, ensure_ascii=False, indent=2)}"
                    if params: req_detail += f"\nParams: {json.dumps(params, ensure_ascii=False, indent=2)}"
                    if body: req_detail += f"\nBody: {json.dumps(body, ensure_ascii=False, indent=2)}"
                    allure.attach(req_detail, "请求报文详情", allure.attachment_type.TEXT)

                    logger.info("⚙️  [发送] %s %s", method, url)
                    response = requests.request(method, url, headers=req_headers, params=params, json=body)
                    res_json = response.json()
                    
                    allure.attach(json.dumps(res_json, ensure_ascii=False, indent=2), "响应报文详情", allure.attachment_type.JSON)
                    logger.debug("📦 [响应] %s...", json.dumps(res_json, ensure_ascii=False)[:150])

                    if case.get('assert_status'):
                        assert response.status_code == int(case['assert_status']), f"HTTP状态码不符"
                    
                    if case.get('assert_jsonpath'):
                        actual_val = self._parse_json_path(res_json, case['assert_jsonpath'])
                        expected_val = case['assert_expected_value']
                        if str(expected_val).lower() == 'true': expected_val = True
                        elif str(expected_val).lower() == 'false': expected_val = False
                        else:
                            try: expected_val = int(expected_val)
                            except: pass
                        assert actual_val == expected_val, f"业务断言失败: 路径 {case['assert_jsonpath']}, 预期 {expected_val}, 实际 {actual_val}"
                    
                    logger.info("✅ [结果] PASSED")
                    passed += 1

            except Exception as e:
                logger.error("❌ [结果] FAILED: %s", str(e))
                failed += 1

        wb.close()
        print(f"\n🚀 [引擎报告] 总计: {passed + failed}, 通过: {passed}, 失败: {failed}")
        return passed, failed